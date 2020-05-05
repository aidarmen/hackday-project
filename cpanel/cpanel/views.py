from django.shortcuts import render
import threading
from google.cloud import firestore
# from google.cloud import storage
import openpyxl
import pyrebase
import uuid
import collections
from  .model.Delivery import Delivery
from .model.User import User
from django.views.decorators.csrf import csrf_protect




# storage_client = storage.Client()

# storage_client = storage.Client.from_service_account_json(
#         '/Users/aidar/Downloads/Chocomarket-hackaton-18f88dd7e4a1.json'
# )


# buckets = list(storage_client.list_buckets())
# print(buckets)

dbfirestore = firestore.Client()

config = {
  'apiKey': "AIzaSyCC8mabGyl5eK4WUpMoXw067ZpvipJuwf0",
  'authDomain': "chocomarket-hackaton-b9628.firebaseapp.com",
  'databaseURL': "https://chocomarket-hackaton-b9628.firebaseio.com",
  'projectId': "chocomarket-hackaton-b9628",
  'storageBucket': "chocomarket-hackaton-b9628.appspot.com",
  'messagingSenderId': "1062233292930",
  'appId': "1:1062233292930:web:828a24518773775ff7749b",
  'measurementId': "G-QJQZ564RT1" }

firebase = pyrebase.initialize_app(config)


# auth = firebase.auth()

def singIn(request):

    return render(request, "signIn.html")

# def postsign(request):
#     email=request.POST.get('email')
#     passw = request.POST.get("pass")
#     try:
#         user = auth.sign_in_with_email_and_password(email,passw)
#     except:
#         message = "invalid cerediantials"
#         return render(request,"signIn.html",{"msg":message})
#     print(user)
#     return render(request, "welcome.html",{"e":email})

def login(request):
    if "GET" == request.method:
        return render(request, 'signIn.html', {})

    else:
        # db = firestore.Client()

        email=request.POST.get('email')
        passw = request.POST.get("pass")
        
        auth = firebase.auth()
        try:
            user = auth.sign_in_with_email_and_password(email, passw)
        except Exception as e:
            message = "invalid cerediantials"
            return render(request,"signIn.html",{"msg":message})

        dbfirestore = firestore.Client()
        docs = dbfirestore.collection(u'users').where(u'email', u'==', email).stream()

            
        if bool(docs):
            for doc in docs:
                request.session['name'] = doc.to_dict().get('name')
        # request.session['uid'] = doc.to_dict().get('uid')
        
        request.session['email'] = user['email']
        request.session['uid'] = user['localId']

        return render(request, 'excel_file.html', {'user': user['email']})
    

def products(request):
    if "GET" == request.method:
        return render(request, 'excel_file.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)
        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        # print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        return render(request, 'excel_file.html', {"excel_data":excel_data})



def view_applications(request):

    dbfirestore = firestore.Client()
    # [START get_simple_query]
    docs = dbfirestore.collection(u'Deliveries').where(u'managerId', u'==', request.session['uid']).stream()

    dlist=[]
    for doc in docs:
        dictlist = doc.to_dict()
        dictlist = collections.OrderedDict(sorted(dictlist.items()))
        dictlist['DeliveryID'] = doc.id
        dlist.append(dictlist)
    
    
    return render(request, 'view_applications.html',{"deliveries":dlist})


def edit_application(request):
    
    idDelivery = request.POST.get('deliveryId')

    dbfirestore = firestore.Client()

    docs = dbfirestore.collection(u'Deliveries').stream()

    deliveryDict = {}
    for doc in docs:
        # print(u'{} => {}'.format(doc.id, doc.to_dict()))
        if (doc.id == idDelivery):
            deliveryDict = doc.to_dict()

    if (bool(deliveryDict)):
        print(deliveryDict)
        request.session['listOfProducts'] = deliveryDict['products']


    
    return render(request, 'edit_application.html',{'idDelivery':idDelivery,'listOfProducts':request.session['listOfProducts']})

# from string to dictionary edit_application returns dictinary in string format needed to reverse
def fromStringToDict(s):
    ch='"'
    arr = [i for i, ltr in enumerate(s) if ltr == ch]
    listtest=[]
    for a in range(0,len(arr),2):
        listtest.append(s[arr[a]+1:arr[a+1]])
        
    tr = dict(zip(listtest[1::4], listtest[3::4]))
    d = {}
    for key in tr:
        d[key] = [tr[key] ]
    return d




def edit_firebase(request):

    content = request.POST.get('dict')
    idDelivery = request.POST.get('idDelivery')
    print(idDelivery)
    # [{"name":"Untitled","quantity":"1"},{"name":"Untitled","quantity":"2"},{"name":"Untitled","quantity":"3"}]
    dicta = fromStringToDict(content)
    
 




    dbfirestore = firestore.Client()
    # [START update_server_timestamp]
    docs = dbfirestore.collection(u'Deliveries').document(idDelivery)
    # deliveryDict = {}
    # for doc in docs:
    #     # print(u'{} => {}'.format(doc.id, doc.to_dict()))
    #     if (doc.id == idDelivery):
    #         deliveryDict = doc.to_dict()
    
    docs.update({
        u'products': dicta
    })

    return render(request, 'success.html',{'msg':'success'})


def chat(request):
    idDelivery = request.POST.get('idDelivery')

    dbfirestore = firestore.Client()

    docs = dbfirestore.collection(u'Deliveries').stream()

    deliveryDict = {}
    for doc in docs:
        # print(u'{} => {}'.format(doc.id, doc.to_dict()))
        if (doc.id == idDelivery):
            deliveryDict = doc.to_dict()

    print(idDelivery)
    if bool(deliveryDict):
        request.session['managerName'] = deliveryDict['managerName']
        request.session['managerId'] = deliveryDict['managerId']
    
    managerName = request.session['managerName']
    managerId = request.session['managerId']
    

    # print(request.POST.get('idDelivery'))
    return render(request, 'chat.html',{"managerName":managerName,"managerId":managerId, "idApplication": idDelivery })

    

def register(request):
    if "GET" == request.method:
        return render(request, 'signUp.html', {})
    else:
        # db = firestore.Client()
        auth = firebase.auth()


        

        email = request.POST.get('email')
        name = request.POST.get('name')
        phone = request.POST.get('phone')
        passw = request.POST.get("pass")
        position = request.POST.get("position")


        user = auth.create_user_with_email_and_password(email, passw)
       

        uid = user['localId']


        newUser = User(email = email ,name= name,phone=phone,password=passw,uid=uid,position=position)
        
# firebase realtime database
        db = firebase.database()
        db.child("users").push(newUser.to_dict(), user['idToken'])

# # firestore cloud database        
        
        dbfirestore.collection(u'users').add(newUser.to_dict())


        request.session['email'] = email
        request.session['name'] = name
        request.session['uid'] = uid
            

        return render(request, 'excel_file.html', {'user': name})
        





def upload_to_firebase(request):
    if "GET" == request.method:
        return render(request, 'excel_file.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)
        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        # print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        excel_dict = {a[0]:a[1:] for a in excel_data}
        del excel_dict['products']


        db = firestore.Client()



        # [START add_custom_class_generated_id]
        delivery = Delivery(products = excel_dict, managerId = request.session['uid'],managerName = request.session['name'] )
        
        db.collection(u'Deliveries').add(delivery.to_dict())

        return render(request, 'success.html', {"msg":"delivery added"})
